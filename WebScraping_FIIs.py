from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from openpyxl import Workbook
import time
from time import ctime
import os

# Obtendo o hórario do computador.

t_inicial = time.time()

# Definindo que o browser utilizado rode em segundo plano.

option = Options()
option.headless = True

# Função para armazenar os dados obtidos dos FIIs em um dicionário.

def dic_fii(lista):
    return {"FII": lista[-1], "Preço": lista[0], "DY12": lista[1], "DY1": lista[2], "P/PV": lista[3], "Liquidez (R$)":
            lista[4], "Valorização": lista[5], "IFIX": lista[6], "Segmento": lista[7]}


# Definindo o browser e abrindo o site de onde serão retirados os nomes dos FIIs.

browser = webdriver.Firefox(options=option, executable_path=r'geckodriver.exe')
browser.get("https://fiis.com.br/lista-de-fundos-imobiliarios/")

# Espera implícita inserida para permitir que todos os componentes do site carreguem.

browser.implicitly_wait(20)

# Obtendo a quantidade de FIIs listados.

qtdd_fiis = browser.find_element_by_xpath('//*[@id="fiis-counter"]').text
int_fiis = int(qtdd_fiis[0:3])

# Lista que armazenará o nome dos FIIs.

lista_FIIs = []

# Através de um loop com range baseado no número de FIIs listados, obtendo os nomes dos FIIs listados no site.

for i in range(int_fiis):
    nome = browser.find_element_by_xpath(f'//*[@id="items-wrapper"]/div[{i+1}]/a/span[1]').text
    lista_FIIs.append(nome)

# Verificação.

print(lista_FIIs)

# Lista que armazenará todos os dados dos FIIs.

todos_FIIs = []

"""
Obtendo todos os dados dos FIIs:
  -Primeiro abre-se o site, concatenando o nome obtido anteriormente com a url base do site.
  -Em sequência há uma verificação, que age caso o site daquele FII em específico não esteja no ar ignorando-o e 
  seguindo para a próxima iteração.
  -Os dados são coletados a partir do XPATH de cada informação.
  -Os dados individuais são armazenados com a função dic_fii em um dicionário e cada dicionário gerado é armazenado na
  lista todos_FIIs.
"""

for fii in lista_FIIs:
    dados_fii = []
    browser.get(f"https://statusinvest.com.br/fundos-imobiliarios/{fii}")
    try:
        element = WebDriverWait(browser, 20).until(ec.presence_of_element_located(
            (By.XPATH, '//*[@id="main-2"]/div[2]/div[1]/div[1]/div/div[1]/strong')))
    except TimeoutException as ex:
        continue
    preco = browser.find_element_by_xpath('//*[@title="Valor atual do ativo"]//*[@class="value"]').text
    dy12 = browser.find_element_by_xpath('//*[@title="Dividend Yield com base nos últimos 12 meses"]'
                                         '//*[@class="value"]').text
    dy1 = browser.find_element_by_xpath('//*[@id="dy-info"]/div/div[2]/div[1]/div[1]/div/b').text
    pvp = browser.find_element_by_xpath('//*[@id="main-2"]/div[2]/div[5]/div/div[2]/div/div[1]/strong').text
    liq = browser.find_element_by_xpath('//*[@id="main-2"]/div[2]/div[6]/div/div/div[3]/div/div/div/strong').text
    val = browser.find_element_by_xpath('//*[@id="main-2"]/div[2]/div[1]/div[5]/div/div[1]/strong').text
    ifix = browser.find_element_by_xpath('//*[@id="main-2"]/div[2]/div[6]/div/div/div[4]/div/a/div/div/strong').text
    segmento = browser.find_element_by_xpath('//*[@id="fund-section"]/div/div/div[2]/div/div[6]/div/div/strong').text
    dados_fii.extend((preco, dy12, dy1, pvp, liq, val, ifix, segmento, fii))
    todos_FIIs.append(dic_fii(dados_fii))

# Verificação.

print(todos_FIIs)

# Gerando o arquivo de Excel que armazenará os dados.

FII_wb = Workbook()
FIIs = FII_wb.active
FIIs.title = "FIIs"

# A data é disposta no canto superior esquerdo da planilha.

data = ctime()
data_celula = FIIs.cell(row=1, column=1)
data_celula.value = data

# Os títulos de cada informação coletada são inseridos na planilha.

for k, chave in enumerate(todos_FIIs[0].keys()):
    titulo = FIIs.cell(row=3, column=k + 2)
    titulo.value = chave

# Por fim os dados são adicionados em fileiras na planilha.

for i, dic_fiis in enumerate(todos_FIIs):
    for j, chave in enumerate(todos_FIIs[0].keys()):
        celula = FIIs.cell(row=i + 4, column=j + 2)
        celula.value = todos_FIIs[i][chave]

# Salvando o arquivo no diretório desejado.

FII_wb.save(os.path.join(r"C:\Users\André\Desktop", "FIIs.xlsx"))

# Encerrando o browser.

browser.quit()

# Tempo de Execução do programa.

print("O programa demorou %s minutos para rodar" % int((time.time() - t_inicial)/60))